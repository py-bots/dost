from ctypes import Union
import os
from pathlib import WindowsPath
from helpers import dostify
output_folder_path = os.path.join(os.path.abspath(
    r'C:\Users\Public\PyBots'), 'My-DOST', 'Excel Folder')

# create output folder if not present
if not os.path.exists(output_folder_path):
    os.makedirs(output_folder_path)


def authenticate_google_spreadsheet(credential_file_path:WindowsPath):
    # import section
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials


    # Logic section
    if not credential_file_path:
        raise Exception("credential (json) file path cannot be empty")

    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
                "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

    creds = ServiceAccountCredentials.from_json_keyfile_name(
        credential_file_path, scope)

    gc = gspread.authorize(creds)

    return gc

def excel_get_dataframe_from_google_spreadsheet(auth, spreadsheet_url:str, sheet_name:str="Sheet1"):
    # import section
    from my_dost.CrashHandler import report_error
    import pandas as pd

    # Logic section

    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    # check if sheet_name is already present in worksheet_list
    sheet_present = False
    for worksheet in worksheet_list:
        if worksheet.title == sheet_name:
            sheet_present = True
            break

    if not sheet_present:
        raise Exception("Sheet name not found")
    else:
        worksheet = sh.worksheet(sheet_name)

    data_frame = pd.DataFrame(worksheet.get_all_records())

def excel_tabular_data_from_website(website_url:str, table_number:int=1):
    """
    Description:
        Gets Website Table Data Easily as an Excel using Pandas. Just pass the URL of Website having HTML Tables.
    Args:
        website_url (str, optional): URL of Website. Defaults to "".
        table_number (int, optional): Table Number. Defaults to all.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (object): Dataframe object.
    """

    # Import Section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    # Logic Section
    
    if not website_url:
        raise Exception("Website URL cannot be empty")

    all_tables = pd.read_html(website_url)

    if not table_number:
        data = all_tables
    else:
        if table_number > len(all_tables):
            raise Exception(
                "Table number cannot be greater than number of tables")

        if table_number < 1:
            raise Exception("Table number cannot be less than 1")

        data = all_tables[table_number - 1]

    return data

def excel_upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url:str, sheet_name:str, df=""):

    # import section
    from my_dost.CrashHandler import report_error
    from gspread_dataframe import set_with_dataframe
    import pandas as pd


    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    if not isinstance(df, pd.DataFrame):
        raise Exception("dataframe must be a pandas dataframe")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    # check if sheet_name is already present in worksheet_list
    sheet_present = False
    for worksheet in worksheet_list:
        if worksheet.title == sheet_name:
            sheet_present = True
            break

    if sheet_present:
        # append df to existing sheet
        worksheet = sh.worksheet(sheet_name)
        row_count = worksheet.get_all_values().__len__()

        if row_count == 0:
            set_with_dataframe(worksheet, dataframe=df)
        else:
            set_with_dataframe(worksheet, dataframe=df,
                                row=row_count + 1, include_column_header=False)

    else:
        worksheet = sh.add_worksheet(
            title=sheet_name, rows="999", cols="26")
        set_with_dataframe(worksheet, df)

def excel_create_file(output_folder:WindowsPath, output_filename:str, output_sheetname:str="Sheet1"):

    # Import section
    from my_dost.CrashHandler import report_error
    import os
    from pathlib import Path
    from openpyxl import Workbook


    # Logic section
    if not output_folder:
        raise Exception("Excel File name cannot be empty")

    if not output_filename:
        raise Exception("Excel File Name cannot be empty")

    if not os.path.exists(output_folder):
        output_folder = output_folder_path

    if ".xlsx" not in output_filename:
        output_filename = os.path.join(output_folder, str(
            Path(output_filename).stem) + ".xlsx")
    else:
        output_filename = os.path.join(output_folder, output_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = output_sheetname

    wb.save(filename=output_filename)

def valid_data(input_filepath:WindowsPath, input_sheetname: str, validate_filepath:bool=True, validate_sheetname:bool=True):

    import os
    from openpyxl import load_workbook
    from my_dost.CrashHandler import report_error

    input_filepath = str(input_filepath)
    input_sheetname = str(input_sheetname)
    if validate_filepath:
        if not ".xlsx" in input_filepath:
            raise Exception(
                "Please provide the excel file name with .xlsx extension")
            return False
        if not os.path.exists(input_filepath):
            raise Exception(
                "Please provide the excel file name with correct path")
            return False
        if validate_sheetname:
            wb = load_workbook(input_filepath)
            sheet_names = wb.sheetnames
            if input_sheetname not in sheet_names:
                raise Exception(
                    "Please provide the correct sheet name")
                print('Available Sheet Names', sheet_names)

def excel_to_dataframe(input_filepath:WindowsPath, input_sheetname:str, header:int=1):

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error



    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not input_sheetname:
        raise Exception("Please provide the sheet name")

    if not valid_data(input_filepath, input_sheetname):
        return [status]

    if header > 0:
        data = pd.read_excel(
            input_filepath, sheet_name=input_sheetname, header=header-1, engine='openpyxl')

    # If the function returns a value, it should be assigned to the data variable.
    # data = value
    return data

def excel_get_row_column_count(df):

    # Description:
    """
    Description:
        Returns the row and column count of the dataframe.


    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): [row_count, column_count] 
    """

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    row, col = df.shape
    row = row + 1
    data = [row, col]

    # If the function returns a value, it should be assigned to the data variable.
    # data = value
    return data

def dataframe_to_excel(df, output_folder:WindowsPath, output_filename:str, output_sheetname:str="Sheet1", mode:str='a'):  # append / overwrite

    # import section
    import pandas as pd
    import os
    from pathlib import Path
    from my_dost.CrashHandler import report_error


    if not output_folder:
        output_folder = output_folder_path
    if not output_filename:
        output_filename = "excel_file"

    if not os.path.exists(output_folder):
        output_folder = output_folder_path

    if ".xlsx" not in output_filename:
        output_filepath = os.path.join(output_folder, str(
            Path(output_filename).stem) + ".xlsx")
    else:
        output_filepath = os.path.join(output_folder, str(
            output_filename))

    if not output_sheetname:
        raise Exception("Please provide the sheet name")

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    new_file = False
    if not os.path.exists(output_filepath):
        # excel_create_file(output_folder, output_filename)
        new_file = True

    if mode == 'a' and not new_file:
        with pd.ExcelWriter(output_filepath, mode="a", engine="openpyxl", if_sheet_exists="overlay",) as writer:
            current_df = excel_to_dataframe(
                output_filepath, output_sheetname)[1]
            row_count = excel_get_row_column_count(current_df)[1]
            df.to_excel(writer, sheet_name=output_sheetname,
                        index=False, startrow=int(row_count[0]), header=False)
    else:
        with pd.ExcelWriter(output_filepath, engine="openpyxl",) as writer:
            df.to_excel(writer, sheet_name=output_sheetname, index=False)

def excel_set_single_cell(df, column_name:str, cell_number:int, text:str):
    """
    Description:
        Writes the given text to the desired column/cell number for the given excel file
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        text (str, optional): Text to be written to the excel file. Defaults to "".

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (df): Modified dataframe
    """

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not text:
        raise Exception("Please provide the text to be set")

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    df.at[cell_number-1, column_name] = text
    data = df

    return data

def excel_get_single_cell(df, column_name:str, cell_number:int,header:int=1):

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not isinstance(column_name, list):
        column_name = [column_name]

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    data = df.at[cell_number-header-1, column_name[0]]

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    return data

def excel_get_all_header_columns(df):

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error
    # Response section
    error = None
    # status = False
    data = None

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    data = df.columns.values.tolist()

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    return data

def excel_get_all_sheet_names(input_filepath:WindowsPath):

    # import section
    from openpyxl import load_workbook
    from my_dost.CrashHandler import report_error


    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not valid_data(input_filepath, validate_sheetname=False):
        raise Exception("Please provide the valid excel path")

    wb = load_workbook(input_filepath)
    data = wb.sheetnames

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    return data

def excel_drop_columns(df, cols:Union[str, list(str)]):

    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not cols:
        raise Exception(
            "Please provide the column name to be dropped.")

    if not isinstance(cols, list):
        cols = [cols]

    df.drop(cols, axis=1, inplace=True)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    return df

def excel_clear_sheet(df):


    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    # Clears the contents of the sheet
    df.drop(df.index, inplace=True)

    data = df

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    return data

def excel_remove_duplicates(df, column_name:str):


    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

    which_one_to_keep = "first"

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        df.drop_duplicates(keep=which_one_to_keep, inplace=True)

    else:
        if not isinstance(column_name, list):
            column_name = [column_name]
        df.drop_duplicates(subset=column_name,
                            keep=which_one_to_keep, inplace=True)

    data = df
    return data

def isNaN(value:str):

    

    # import section
    from my_dost.CrashHandler import report_error

    # Response section
    error = None
    # status = False

    if not value:
        raise Exception(
            "Value is empty. Please give a value to check.")
    import math
    status = math.isnan(float(value))

def df_from_list(list_of_lists:list, column_names:list(str)):
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    # Logic section
    if not isinstance(list_of_lists, list):
        raise Exception("Please pass input as list of lists")

    if isinstance(list_of_lists, list):
        if column_names == None:
            data = pd.DataFrame(list_of_lists)
        else:
            data = pd.DataFrame(list_of_lists, columns=column_names)

    return data

def df_from_string(df_string: str, word_delimeter:str=" ", line_delimeter:str="\n", column_names:list=None):
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error
    # Logic section
    if not df_string:
        raise Exception("Please pass input as string")

    if not isinstance(df_string, str):
        df_string = str(df_string)

    if column_names == None:
        data = pd.DataFrame([x.split(word_delimeter)
                            for x in df_string.split(line_delimeter)])
    elif isinstance(column_names, list):
        data = pd.DataFrame([x.split(word_delimeter) for x in df_string.split(
            line_delimeter)], columns=column_names)
    return data

def df_extract_sub_df(df, row_start: int, row_end: int, column_start: int, column_end: int):
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    # Logic section
    # try:
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        data = df.iloc[row_start:row_end, column_start:column_end]

    return data

def set_value_in_df(df, row_number: int, column_number: int, value:str):
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

    # Logic section
    # try:
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    # print(type(df))
    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        df.iloc[row_number-1, column_number-1] = value

    # except Exception as ex:
    #     report_error(ex)
    #     error = ex

    # # else:
    # #     status = True

    # finally:
    #     if error is not None:
    #         raise Exception(error)
        return df

def get_value_in_df(df, row_number: int, column_number: int):
   
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error


    # Logic section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        data = df.iloc[row_number-1, column_number-1]

    return data

def df_drop_rows(df, row_start: int, row_end: int):
   
    # import section
    import pandas as pd
    from my_dost.CrashHandler import report_error

   
    # Logic section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        # -1 because index starts from 0
        data = df.drop(df.index[row_start-1:row_end])

    return data