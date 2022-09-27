"""
Excel Module for my_dost. This module contains functions for working with excel and spreadsheets

Examples:
    >>> excel_get_row_column_count(df)
        (10, 5)
    >>> excel_get_single_cell("C:\\Users\\user\\Desktop\\excel_file.xlsx", "Column1", 1)
        "abc"
    >>> excel_create_file(output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1")

This module contains the following functions:

- `authenticate_google_spreadsheet(credential_file_path)`: This creates authentication object for google spreadsheet.
- `excel_get_dataframe_from_google_spreadsheet(auth, spreadsheet_url, sheet_name)`: Get dataframe from google spreadsheet.
- `excel_tabular_data_from_website(url, table_id)`: Get tabular data from website.
- `excel_upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url, sheet_name, df)`: Upload dataframe to google spreadsheet.
- `excel_create_file(output_folder, output_filename, output_sheetname)`: Create excel file.
- `valid_data(input filepath, input sheet name, validate filepath, validate sheet name)`: Check if data is valid.
- `excel_to_dataframe(excel_file_path, sheet_name,header)`: Convert excel file to dataframe.
- `excel_get_row_column_count(df)`: Get row and column count of dataframe.
- `dataframe_to_excel(df, excel_folder, excel_file_name, sheet_name, mode)`: Convert dataframe to excel file.
- `excel_set_single_cell(df, column_name, cell_number, value)`: Set single cell value in excel file.
- `excel_get_single_cell(df, column_name, cell_number, header)`: Get single cell value from excel file.
- `excel_get_all_header_columns(df)`: Get all header columns from excel file.
- `excel_get_all_sheet_names(excel_file_path)`: Get all sheet names from excel file.
- `excel_drop_columns(dataframe, columns_to_drop)`: Drop columns from data frame.
- `excel_clear_sheet(dataframe)`: Clear sheet from excel file.
- `excel_remove_duplicates(dataframe, column_name)`: Remove duplicates from excel file.
- `isNaN(value)`: Check if value is NaN.
- `df_from_list(list_of_lists, header names)`: Create dataframe from list of lists.
- `df_from_string(string, delimiter)`: Create dataframe from string.
- `df_extract_sub_df(dataframe, row start, row end, column start, column end)`: Extract sub dataframe from dataframe.
- `set_value_in_df(dataframe, column_name, row_number, value)`: Set value in dataframe.
- `get_value_in_df(dataframe, column_name, row_number)`: Get value from dataframe.
- `df_drop_rows(dataframe, row start, row end)`: Drop rows from dataframe.
"""


from ctypes import Union
from multiprocessing.sharedctypes import Value
import pandas as pd
import os
from pathlib import WindowsPath
from helpers import dostify
from typing import List
output_folder_path = os.path.join(os.path.abspath(
    r'C:\Users\Public\PyBots'), 'My-DOST', 'Excel Folder')

# create output folder if not present
if not os.path.exists(output_folder_path):
    os.makedirs(output_folder_path)

@dostify(errors=[])
def authenticate_google_spreadsheet(credential_file_path:WindowsPath):
    """This creates
    
    Args:
        credential_file_path (WindowsPath): Path to JSON file containing credentials
    
    Returns:
    
    Examples:
        >>> authenticate_google_spreadsheet("C:\Users\Asus\Downloads\Credential_file.json")
    """


    # Import Section
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials


    # Code section
    if not credential_file_path:
        raise Exception("credential (json) file path cannot be empty")

    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
                "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

    creds = ServiceAccountCredentials.from_json_keyfile_name(
        credential_file_path, scope)

    gc = gspread.authorize(creds)

    return gc

@dostify(errors=[])
def excel_get_dataframe_from_google_spreadsheet(auth, spreadsheet_url:str, sheet_name:str="Sheet1") -> pd.DataFrame:
    """ Get dataframe from google spreadsheet

    Args:
        auth (object): Authentication object.
        spreadsheet_url (str): Spreadsheet URL.
        sheet_name (str): Sheet name.
    
    Returns:
        df (pd.DataFrame): Dataframe object.

    Examples:
        >>> excel_get_dataframe_from_google_spreadsheet(auth,spreadsheet_url="https://docs.google.com/spreadsheets/d/1X2X3X4X5X6X7X8X9X/edit#gid=0", sheet_name="Sheet1")
        df
    """


    # import section
    import pandas as pd

    # Code section
    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    # check if sheet_name is already present in worksheet_list
    sheet_present = False
    for worksheet in worksheet_list:
        if worksheet.title == sheet_name:
            sheet_present = True
            break

    if not sheet_present:
        raise Exception("Sheet name not found")
    else:
        worksheet = sh.worksheet(sheet_name)

    data_frame = pd.DataFrame(worksheet.get_all_records())

@dostify(errors=[])
def excel_tabular_data_from_website(website_url:str, table_number:int=1) -> pd.DataFrame:
    """Returns a dataframe from a website table.
   
    Args:
       website_url (str): Website URL.
       table_number (int, optional): Table number. Defaults to 1.
    
    Examples:
        >>> excel_tabular_data_from_website(website_url="https://www.worldometers.info/coronavirus/", table_number=1)
    """

    # Import Section
    import pandas as pd
    
    # Code Section
    if not website_url:
        raise Exception("Website URL cannot be empty")

    all_tables = pd.read_html(website_url)

    if not table_number:
        data = all_tables
    else:
        if table_number > len(all_tables):
            raise Exception(
                "Table number cannot be greater than number of tables")

        if table_number < 1:
            raise Exception("Table number cannot be less than 1")

        data = all_tables[table_number - 1]

    return data

@dostify(errors=[])
def excel_upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url:str, sheet_name:str, df:pd.DataFrame) -> None:
    """Uploads a dataframe to a google spreadsheet.
    
    Args:
        auth (object): Authentication object.
        spreadsheet_url (str): Spreadsheet URL.
        sheet_name (str): Sheet name.
        df (pd.DataFrame): Dataframe object.
    
    Examples:
        >>> excel_upload_dataframe_to_google_spreadsheet(auth=auth, spreadsheet_url="https://docs.google.com/spreadsheets/d/1X2X3X4X5X6X7X8X9X/edit#gid=0", sheet_name="Sheet1", df=df)
    """


    # import section
    from gspread_dataframe import set_with_dataframe
    import pandas as pd


    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    if not isinstance(df, pd.DataFrame):
        raise Exception("dataframe must be a pandas dataframe")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    # check if sheet_name is already present in worksheet_list
    sheet_present = False
    for worksheet in worksheet_list:
        if worksheet.title == sheet_name:
            sheet_present = True
            break

    if sheet_present:
        # append df to existing sheet
        worksheet = sh.worksheet(sheet_name)
        row_count = worksheet.get_all_values().__len__()

        if row_count == 0:
            set_with_dataframe(worksheet, dataframe=df)
        else:
            set_with_dataframe(worksheet, dataframe=df,
                                row=row_count + 1, include_column_header=False)

    else:
        worksheet = sh.add_worksheet(
            title=sheet_name, rows="999", cols="26")
        set_with_dataframe(worksheet, df)

@dostify(errors=[])
def excel_create_file(output_folder:WindowsPath, output_filename:str, output_sheetname:str="Sheet1") -> None:
    """ Creates an excel file with a sheet in the specified folder.
    
    Args:
        output_folder (WindowsPath): Output folder path.
        output_filename (str): Output file name.
        output_sheetname (str, optional): Output sheet name. Defaults to "Sheet1".
    
    Examples:
        >>> excel_create_file(output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1")
    """
    
    
    # Import Section
    import os
    from pathlib import Path
    from openpyxl import Workbook


    # Code Section
    if not output_folder:
        raise Exception("Excel File name cannot be empty")

    if not output_filename:
        raise Exception("Excel File Name cannot be empty")

    if not os.path.exists(output_folder):
        output_folder = output_folder_path

    if ".xlsx" not in output_filename:
        output_filename = os.path.join(output_folder, str(
            Path(output_filename).stem) + ".xlsx")
    else:
        output_filename = os.path.join(output_folder, output_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = output_sheetname

    wb.save(filename=output_filename)

@dostify(errors=[])
def valid_data(input_filepath:WindowsPath, input_sheetname: str, validate_filepath:bool=True, validate_sheetname:bool=True) -> bool:
    """This function validates the input file path and sheet name.

    Args:
        input_filepath (WindowsPath): Input file path.
        input_sheetname (str): Input sheet name.
        validate_filepath (bool, optional): Whether to validate file path or not. Defaults to True.
        validate_sheetname (bool, optional): Whether to validate sheet name or not. Defaults to True.

    Returns:
        bool: True if valid, False if invalid.
    
    Examples:
        >>> valid_data(input_filepath="C:\\Users\\user\\Desktop\\test.xlsx", input_sheetname="Sheet1")
        True
    """
    # Import Section
    import os
    from openpyxl import load_workbook
    
    # Code Section
    input_filepath = str(input_filepath)
    input_sheetname = str(input_sheetname)
    if validate_filepath:
        if not ".xlsx" in input_filepath:
            raise Exception(
                "Please provide the excel file name with .xlsx extension")
            return False
        if not os.path.exists(input_filepath):
            raise Exception(
                "Please provide the excel file name with correct path")
            return False
        if validate_sheetname:
            wb = load_workbook(input_filepath)
            sheet_names = wb.sheetnames
            if input_sheetname not in sheet_names:
                raise Exception(
                    "Please provide the correct sheet name")
        return True

@dostify(errors=[])
def excel_to_dataframe(input_filepath:Union[str,WindowsPath], input_sheetname:str, header:int=1) -> pd.DataFrame:
    """Converts excel file to dataframe.
    
    Args:
        input_filepath (Union[str,WindowsPath]): Input file path.
        input_sheetname (str): Input sheet name.
        header (int, optional): Header row number. Defaults to 1.
    
    Returns:
        pd.DataFrame: Dataframe of the excel file.
        
    Examples:
        >>> excel_to_dataframe(input_filepath="C:\\Users\\user\\Desktop\\test.xlsx", input_sheetname="Sheet1")
        dataframe
    """


    # Import Section
    import pandas as pd
    
    # Code Section
    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not input_sheetname:
        raise Exception("Please provide the sheet name")

    if not valid_data(input_filepath, input_sheetname):
        raise ValueError("File does not contain valid data")

    if header > 0:
        data = pd.read_excel(
            input_filepath, sheet_name=input_sheetname, header=header-1, engine='openpyxl')
    return data

@dostify(errors=[])
def excel_get_row_column_count(df) -> tuple(int,int):
    """ Returns the row and column count of the dataframe
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
    
    Returns:
        tuple: Row and column count of the dataframe.
        
    Examples:
        >>> excel_get_row_column_count(df)
        (10, 5)
    """

    
    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    row, col = df.shape
    row = row + 1
    data = (row, col)

    # If the function returns a value, it should be assigned to the data variable.
    # data = value
    return data

@dostify(errors=[])
def dataframe_to_excel(df, output_folder:WindowsPath, output_filename:str, output_sheetname:str="Sheet1", mode:str='a') -> None:  # append / overwrite
    """ Converts the dataframe to excel file
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        output_folder (WindowsPath): Output folder path.
        output_filename (str): Output file name.
        output_sheetname (str, optional): Output sheet name. Defaults to "Sheet1".
        mode (str, optional): Mode of the excel file. Defaults to 'a'.
        
    Examples:
        >>> dataframe_to_excel(df, output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1", mode='a')
    """

    # import section
    import pandas as pd
    import os
    from pathlib import Path
    
    # Code Section
    if not output_folder:
        output_folder = output_folder_path
    if not output_filename:
        output_filename = "excel_file"

    if not os.path.exists(output_folder):
        output_folder = output_folder_path

    if ".xlsx" not in output_filename:
        output_filepath = os.path.join(output_folder, str(
            Path(output_filename).stem) + ".xlsx")
    else:
        output_filepath = os.path.join(output_folder, str(
            output_filename))

    if not output_sheetname:
        raise Exception("Please provide the sheet name")

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    new_file = False
    if not os.path.exists(output_filepath):
        # excel_create_file(output_folder, output_filename)
        new_file = True

    if mode == 'a' and not new_file:
        with pd.ExcelWriter(output_filepath, mode="a", engine="openpyxl", if_sheet_exists="overlay",) as writer:
            current_df = excel_to_dataframe(
                output_filepath, output_sheetname)[1]
            row_count = excel_get_row_column_count(current_df)[1]
            df.to_excel(writer, sheet_name=output_sheetname,
                        index=False, startrow=int(row_count[0]), header=False)
    else:
        with pd.ExcelWriter(output_filepath, engine="openpyxl",) as writer:
            df.to_excel(writer, sheet_name=output_sheetname, index=False)

@dostify(errors=[])
def excel_set_single_cell(df, column_name:str, cell_number:int, text:str) -> pd.DataFrame:
    """
    Description:
        Writes the given text to the desired column/cell number for the given excel file
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        text (str, optional): Text to be written to the excel file. Defaults to "".

    Returns:
        data (df): Modified dataframe
    
    Examples:
        >>> excel_set_single_cell(dataframe, "Column1", 1, "abc")
        df
    """

    # import section
    import pandas as pd
    
    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not text:
        raise Exception("Please provide the text to be set")

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    df.at[cell_number-1, column_name] = text
    data = df

    return data

@dostify(errors=[])
def excel_get_single_cell(df, column_name:str, cell_number:int,header:int=1) -> str:
    """Gets the text from the desired column/cell number for the given excel file
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        header (int, optional): Header row number. Defaults to 1.
    
    Returns:
        data (str): Text from the desired column/cell number for the given excel file
    
    Examples:
        >>> excel_get_single_cell("C:\\Users\\user\\Desktop\\excel_file.xlsx", "Column1", 1)
        "abc"
    """

    # Import Section
    import pandas as pd
    
    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not isinstance(column_name, list):
        column_name = [column_name]

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    data = df.at[cell_number-header-1, column_name[0]]

    return data

@dostify(errors=[])
def excel_get_all_header_columns(df) -> List[str]:
    """Gets all header columns from the excel file
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
    
    Returns:
        data (list): List of header columns
    
    Examples:
        >>> excel_get_all_header_columns("C:\\Users\\user\\Desktop\\excel_file.xlsx")
        ["Column1", "Column2"]
    """


    # import section
    import pandas as pd
    
    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    data = df.columns.values.tolist()
    return data

@dostify(errors=[])
def excel_get_all_sheet_names(input_filepath:WindowsPath) -> List[str]:
    """Gets the sheet names from the excel file
    
    Args:
        input_filepath (str): Path of the excel file.
    
    Returns:
        data (list): List of sheet names
        
    Examples:
        >>> excel_get_all_sheet_names("C:\\Users\\user\\Desktop\\excel_file.xlsx")
        ["Sheet1", "Sheet2"]
    """


    # Import Section
    from openpyxl import load_workbook
    
    # Code Section
    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not valid_data(input_filepath, validate_sheetname=False):
        raise Exception("Please provide the valid excel path")

    wb = load_workbook(input_filepath)
    data = wb.sheetnames
    return data

@dostify(errors=[])
def excel_drop_columns(df, cols:Union[str, list(str)]) -> pd.DataFrame:
    """Drops the columns from the excel file
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        cols (str, list(str)): Column name to be dropped.
    
    Returns:
        data (df): Modified dataframe
    
    Examples:
        >>> excel_drop_columns(df, "column_name")
        df
    """


    # Import Section
    import pandas as pd
    
    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not cols:
        raise Exception(
            "Please provide the column name to be dropped.")

    if not isinstance(cols, list):
        cols = [cols]

    df.drop(cols, axis=1, inplace=True)
    return df

@dostify(errors=[])
def excel_clear_sheet(df) -> pd.DataFrame:
    """Clears the sheet
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
    
    Returns:
        data (df): Modified dataframe
    
    Examples:
        >>> excel_clear_sheet(df)
        df
    """


    # Import Section
    import pandas as pd
    
    #Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    # Clears the contents of the sheet
    df.drop(df.index, inplace=True)

    data = df
    return data

@dostify(errors=[])
def excel_remove_duplicates(df, column_name:str) -> pd.DataFrame:
    """Removes the duplicates from the given column
    
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
    
    Returns:
        data (df): Modified dataframe
    
    Examples:
        >>> excel_remove_duplicates(df, "column_name")
        df
    """


    # Import Section
    import pandas as pd
    
    # Code Section
    which_one_to_keep = "first"

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        df.drop_duplicates(keep=which_one_to_keep, inplace=True)

    else:
        if not isinstance(column_name, list):
            column_name = [column_name]
        df.drop_duplicates(subset=column_name,
                            keep=which_one_to_keep, inplace=True)

    data = df
    return data

@dostify(errors=[])
def isNaN(value:str) -> bool:
    """Checks if the value is NaN

    Args:
        value (str): value to be checked

    Returns:
        bool: True if value is NaN, False otherwise
    
    Examples:
        >>> isNaN("abc")
        False
    """

    # Code Section
    if not value:
        raise Exception(
            "Value is empty. Please give a value to check.")
    import math
    return math.isnan(float(value))

@dostify(errors=[])
def df_from_list(list_of_lists:list, column_names:list(str)) -> pd.DataFrame:
    """Converts list of lists to dataframe
    
    Args:
        list_of_lists (list): list of lists to be converted to dataframe
        column_names (list): column names
        
    Returns:
        pandas dataframe: dataframe
        
    Examples:
        >>> df_from_list([[1,2,3],[4,5,6]], ["col1", "col2", "col3"])
        dataframe
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(list_of_lists, list):
        raise Exception("Please pass input as list of lists")

    if isinstance(list_of_lists, list):
        if column_names == None:
            data = pd.DataFrame(list_of_lists)
        else:
            data = pd.DataFrame(list_of_lists, columns=column_names)

    return data

@dostify(errors=[])
def df_from_string(df_string: str, word_delimeter:str=" ", line_delimeter:str="\n", column_names:list=None) -> pd.DataFrame:
    """Converts string to dataframe

    Args:
        df_string (str): string to be converted to dataframe
        word_delimeter (str): word delimeter.Defaults to space
        line_delimeter (str): line delimeter. Defaults to new line
        column_names (list): column names. Defaults to None
    
    Returns:
        pandas dataframe: dataframe
    
    Examples:
        >>> df_from_string("a b c d e f g h i j k l m n o p q r s t u v w x y z", " ", "\n", ["col1", "col2", "col3"])
        dataframe
    
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not df_string:
        raise Exception("Please pass input as string")

    if not isinstance(df_string, str):
        df_string = str(df_string)

    if column_names == None:
        data = pd.DataFrame([x.split(word_delimeter)
                            for x in df_string.split(line_delimeter)])
    elif isinstance(column_names, list):
        data = pd.DataFrame([x.split(word_delimeter) for x in df_string.split(
            line_delimeter)], columns=column_names)
    return data

@dostify(errors=[])
def df_extract_sub_df(df, row_start: int, row_end: int, column_start: int, column_end: int) -> pd.DataFrame:
    """Extracts sub dataframe from the given dataframe
    
    Args:
        df (pandas dataframe): dataframe
        row_start (int): row start (inclusive)
        row_end (int): row end   (exclusive)
        column_start (int): column start (inclusive)
        column_end (int): column end (exclusive)
    
    Returns:
        pandas dataframe(pandas dataframe): sub dataframe
        
    Examples:
        >>> df_extract_sub_df(df, 1, 2, 3, 4)
        sub_dataframe
    
    """
    
    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        data = df.iloc[row_start:row_end, column_start:column_end]

    return data

@dostify(errors=[])
def set_value_in_df(df, row_number: int, column_number: int, value:str) -> pd.DataFrame:
    """Sets value in dataframe
    
    Args:
        df (pandas dataframe): dataframe to be modified
        row_number (int): Row number of the cell
        column_number (int): Column number of the cell
        value (str): value to be set in the cell
        
    Returns:
        pandas dataframe: dataframe with value set
    
    Examples:
        >>> set_value_in_df(df, 1, 2, "abc")
        modified_dataframe
    
    """
    
    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        df.iloc[row_number-1, column_number-1] = value

        return df

@dostify(errors=[])
def get_value_in_df(df, row_number: int, column_number: int) -> str:
    """Gets value from dataframe
    
    Args:
        df (pandas dataframe): dataframe
        row_number (int): Row number of the cell
        column_number (int): Column number of the cell
        
    Returns:
        str: value in the cell
    Examples:
        >>> get_value_in_df(df, 1, 2)
        abc
    
    """
    
    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        data = df.iloc[row_number-1, column_number-1]

    return data

@dostify(errors=[])
def df_drop_rows(df, row_start: int, row_end: int) -> pd.DataFrame: 
    """Drops rows from dataframe
    
    Args:
        df (pandas dataframe): dataframe
        row_start (int): row start (inclusive)
        row_end (int): row end   (exclusive)

    Returns:
        pandas dataframe: dataframe with rows dropped
    Examples:
        >>> df = df_drop_rows(df, 1, 2)
    
    """
   
    # Import Section
    import pandas as pd
   
    # Code section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        # -1 because index starts from 0
        data = df.drop(df.index[row_start-1:row_end])

    return data