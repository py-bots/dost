
def string_extract_substring(string: str, start_word: str, end_word: str, include_start_word: bool=False, include_end_word: bool=False) -> str:
    """Extracts substring from a string.

    Args:
        string ([type]): [description]
        start_word ([type]): [description]
        end_word ([type]): [description]
        include_start_word (bool, optional): [description]. Defaults to False.
        include_end_word (bool, optional): [description]. Defaults to False.

    Returns:
        data (str): Extracted substring.

    Examples:
        >>> string_extract_substring('Hello World', 'Hello', 'World')
    """

    # Code Section
    if (string.find(start_word) == -1) or (string.find(end_word) == -1):
        raise ValueError("Start word or end word not found in string.")
    
    data = string[string.find(start_word)+len(start_word*int(not include_start_word)):string.find(end_word)+len(end_word*int(include_end_word))]
    
    return data



##################################################################################



def date_time_now() -> str:
    """Get current date and time.

    Returns:
        data (str): Current date and time.

    Examples:
        >>> date_time_now()
        '2020-01-01 00:00:00'
    """
    # Import Section
    import datetime

    # Code Section
    return datetime.datetime.now()

def add_datetime(input_date: str, time_type: str, value: int=0) -> str:
    """Add time to a date.

    Args:
        input_date (str): Date to add time to.
        time_type (str): Type of time to add. Can be minutes, hours, days, weeks, months.
        value (int): Value of time to add.
    
    Returns:
        data (str): Date with added time.
    
    Examples:
        >>> add_datetime('2020-01-01', 'days', 1)
        '2020-01-02 00:00:00'
    """
    # Import Section
    import pendulum

    # Code Section
    if not isinstance(input_date, str):
        input_date = str(input_date)
        
    minutes=hours=days=weeks=months=0
    if time_type == 'minutes':
        minutes = value
    elif time_type == 'hours':
        hours = value
    elif time_type == 'days':
        days = value
    elif time_type == 'weeks':
        weeks = value
    elif time_type == 'months':
        months = value
    return pendulum.parse(input_date).add(minutes=minutes, hours=hours, days=days, weeks=weeks, months=months).to_datetime_string()

def subtract_datetime(input_date: str, time_type: str='days', value: int=0) -> str:
    """Subtract time from a date.

    Args:
        input_date (str): Date to subtract time from.
        time_type (str): Type of time to subtract. Can be minutes, hours, days, weeks, months.
        value (int): Value of time to subtract.
    
    Returns:
        data (str): Date with subtracted time.
    
    Examples:
        >>> subtract_datetime('2020-01-01', 'days', 1)
        '2019-12-31 00:00:00'
    """

    # Import Section
    import pendulum

    # Code Section
    if not isinstance(input_date, str):
        input_date = str(input_date)

    minutes=hours=days=weeks=months=0
    if time_type == 'minutes':
        minutes = value
    elif time_type == 'hours':
        hours = value
    elif time_type == 'days':
        days = value
    elif time_type == 'weeks':
        weeks = value
    elif time_type == 'months':
        months = value
    return pendulum.parse(input_date).subtract(minutes=minutes, hours=hours, days=days, weeks=weeks, months=months).to_datetime_string()



##################################################################################




def mouse_move(x: int, y: int, type_of_movement: str="abs") -> None:
    """Moves the cursor to the given X Y Co-ordinates.

    Args:
        x (int): x-coordinate on screen.
        y (int): y-coordinate on screen.
        type_of_movement (str): Type of movement (Absolute or Relative to current Position).Defaults to abs.

    Examples:
        >>> mouse_move(100, 100)
    """

    # import section
    import time
    import pywinauto as pwa
    import win32api

    # code section
    if type_of_movement == "abs" or type_of_movement == "rel":

        if type_of_movement == "abs":
            x, y = int(x), int(y)
            time.sleep(0.1)
            # pg.moveTo(x, y)
            pwa.mouse.move(coords=(x, y))
            time.sleep(0.1)
        elif type_of_movement == "rel":
            current_x, current_y = win32api.GetCursorPos()
            x, y = int(x), int(y)
            current_x, current_y = int(current_x), int(current_y)
            x, y = int(current_x + x), int(current_y + y)
            pwa.mouse.move(coords=(x, y))
            time.sleep(0.1)
        else:
            raise Exception(
                "Please check the 'type of movement' value.")

    else:
        raise Exception("Type of movement is required.")


def mouse_drag_from_to(x1: int, y1: int, x2: int, y2: int) -> None:
    """Drags the mouse from one point to another.

    Args:
        x1 (int): x-coordinate of starting point.
        y1 (int): y-coordinate of starting point.
        x2 (int): x-coordinate of ending point.
        y2 (int): y-coordinate of ending point.
    
    Examples:
        >>> mouse_drag_from_to(100, 100, 200, 200)
    """

    # Import Section
    import time
    import pywinauto as pwa

    # Code Section  
    time.sleep(0.1)
    x1, y1 = int(x1), int(y1)
    x2, y2 = int(x2), int(y2)
    pwa.mouse.move(coords=(x1, y1))
    pwa.mouse.press(coords=(x1, y1))
    pwa.mouse.move(coords=(x2, y2))
    pwa.mouse.release(coords=(x2, y2))
    time.sleep(0.1)



##################################################################################


from typing import Union,List
from pathlib import WindowsPath
def email_send_via_desktop_outlook(to_email_id: Union[str, List[str]], subject: str="", message: str="", attachment_path: Union[str, List[str], WindowsPath, List[WindowsPath]]=""):
    """Sends an email via desktop outlook.

    Args:
        to_email_id (str || list): Email id of the recipient.
        subject (str): Subject of the email.
        message (str): Message of the email.
        attachment_path (str || Windows Path || list): Path of the attachment.
    
    Examples:
        >>> email_send_via_desktop_outlook('abc@gmail.com', 'Subject', 'Message', 'C:\\Users\\abc\\Desktop\\file.txt')
    """

    # Import Section
    from pathlib import Path
    import win32com.client

    # Code Section
    outlook = win32com.client.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)

    if type(to_email_id) is list:
        for m in to_email_id:
            mail.Recipients.Add(m)
    else:
        mail.To = to_email_id

    mail.Subject = subject

    mail.HTMLBody = f"<body><html> {message} <br> </body></html>"

    if attachment_path:
        if isinstance(attachment_path, list):
            for p in attachment_path:
                mail.Attachments.Add(p)
        else:
            mail.Attachments.Add(attachment_path)

    mail.Send()


##################################################################################



def file_get_json_details(path_of_json_file: Union[str, WindowsPath], section: str) -> dict:
    """Returns the details of the given section in the json file.

    Args:
        path_of_json_file (str || Windows Path): Path of the json file.
        section (str): Section of the json file.
    
    Returns:
        data (dict): Details of the given section in the json file.
    
    Examples:
        >>> file_get_json_details('C:\\Users\\abc\\Desktop\\file.json', 'section')
    """

    # Import Section
    import json
    
    # Code Section
    if not path_of_json_file:
        raise Exception("Path of json file is empty")

    if not section:
        raise Exception("Section is empty")

    # import json

    with open(path_of_json_file, 'r') as fp:
        data = json.load(fp)
    fp.close()

    if section in list(data.keys()):
        data = data.get(section)
    else:
        raise Exception(
            'Section can\'t be find in given json file.')
    
    return data





#######################################################################################

import pandas as pd
def df_vlookup(df1: pd.DataFrame, df2: pd.DataFrame, column_name: str, how: str = "left") -> pd.DataFrame:
    """Performs vlookup operation on two dataframes.

    Args:
        df1 (pd.DataFrame): First dataframe.
        df2 (pd.DataFrame): Second dataframe.
        column_name (str): Column name on which vlookup is to be performed.
        how (str): Type of vlookup. Default is 'left'.
    
    Returns:
        data (pd.DataFrame): Dataframe after vlookup operation.
    
    Examples:
        >>> df_vlookup(df1, df2, 'column_name', 'left')
    """

    # Import Section
    import pandas as pd

    # Code Section
    if df1.empty or df2.empty:
        raise Exception("Dataframe cannot be empty")

    if column_name not in df1.columns or column_name not in df2.columns:
        raise Exception("Please pass correct column name")

    if isinstance(df1, pd.DataFrame) and isinstance(df2, pd.DataFrame):
        data = pd.merge(df1, df2, on=column_name, how=how)
    return data




def excel_merge_all_files(input_folder_path: Union[str, WindowsPath], output_folder: Union[str, WindowsPath], output_filename: str) -> None:
    """Merges all the excel files in the given folder.

    Args:
        input_folder_path (str || Windows Path): Path of the input folder.
        output_folder (str || Windows Path): Path of the output folder.
        output_filename (str): Name of the output file.
    
    Examples:
        >>> excel_merge_all_files('C:\\Users\\abc\\Desktop\\input_folder', 'C:\\Users\\abc\\Desktop\\output_folder', 'output_file')
    """

    # Import Section
    import os
    import pandas as pd
    import datetime
  

    # Code Section
    if not input_folder_path:
        raise Exception("Input folder path is empty.")


    if not output_filename:
        time_stamp_now = str(datetime.datetime.now().strftime("%m-%d-%Y"))
        output_filename = "excel_merged" + time_stamp_now

    filelist = [f for f in os.listdir(
        input_folder_path) if f.endswith(".xlsx")]

    all_excel_file_lst = []

    for file1 in filelist:
        file_path = os.path.join(input_folder_path, str(file1))
        file_path = str(file_path)

        all_excel_file = pd.read_excel(
            file_path, dtype=str, engine='openpyxl')
        all_excel_file_lst.append(all_excel_file)

    appended_df = pd.concat(all_excel_file_lst)

    final_path = os.path.join(
        output_folder, output_filename + ".xlsx")
    appended_df.to_excel(final_path, index=False)




def excel_copy_range_from_sheet(input_filepath: Union[str, WindowsPath], input_sheetname: str, start_row: int = 1, start_col: int = 1, end_row: int = 1, end_col: int = 1):
    """Copies the range from the given sheet.

    Args:
        input_filepath (str || Windows Path): Path of the input excel file.
        input_sheetname (str): Name of the input sheet.
        start_row (int): Starting row number. Default is 1.
        start_col (int): Starting column number. Default is 1.
        end_row (int): Ending row number. Default is 1.
        end_col (int): Ending column number. Default is 1.
    
    Returns:
        data (list): List of copied data.
    
    Examples:
        >>> excel_copy_range_from_sheet('C:\\Users\\abc\\Desktop\\file.xlsx', 'sheet_name', 1, 1, 1, 1)
    """

    # Import Section
    from openpyxl import load_workbook


    # Code Section
    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not input_sheetname:
        raise Exception("Please provide the sheet name")

    if start_col == 0 and start_row == 0 and end_col == 0 and end_row == 0:
        raise Exception("Please provide the range to be copied.")

    if not valid_data(input_filepath, input_sheetname):
        raise Exception("Please provide valid data.")

    from_wb = load_workbook(filename=input_filepath)
    try:
        fromSheet = from_wb[input_sheetname]
    except:
        fromSheet = from_wb.worksheets[0]
    rangeSelected = []

    if end_row < start_row:
        end_row = start_row

    # Loops through selected Rows
    for i in range(start_row, end_row + 1, 1):

        rowSelected = []
        for j in range(start_col, end_col+1, 1):
            rowSelected.append(
                fromSheet.cell(row=i, column=j).value)

        rangeSelected.append(rowSelected)

    data = rangeSelected

    return data


def excel_paste_range_to_sheet(input_filepath: Union[str, WindowsPath], input_sheetname: str = 'Sheet 1', start_row: int = 1, start_col: int = 1, copied_data: list =[]):
    """Pastes the copied data to the given sheet.

    Args:
        input_filepath (str || Windows Path): Path of the input excel file.
        input_sheetname (str): Name of the input sheet.
        start_row (int): Starting row number. Default is 1.
        start_col (int): Starting column number. Default is 1.
        copied_data (list): List of copied data.

    Examples:
        >>> excel_paste_range_to_sheet('C:\\Users\\abc\\Desktop\\file.xlsx', 'sheet_name', 1, 1, copied_data)
    """

    # Import Section
    from openpyxl import load_workbook

    # Code Section
    if not copied_data:
        raise Exception(
            "Copied data is empty. First copy the data using Copy Range From Sheet function.")

    if not input_filepath:
        raise Exception("Excel path is empty.")

    if start_col == 0 or start_row == 0:
        raise Exception("Please provide the range to be copied.")

    if not valid_data(input_filepath, input_sheetname):
        return [False, None]

    to_wb = load_workbook(filename=input_filepath)

    toSheet = to_wb[input_sheetname]

    # Get the number of rows and columns in the copied data
    rowCount = len(copied_data)
    colCount = len(copied_data[0])

    endRow = start_row + rowCount - 1
    endCol = start_col + colCount - 1

    countRow = 0
    for i in range(start_row, endRow+1, 1):
        countCol = 0
        for j in range(start_col, endCol+1, 1):
            toSheet.cell(
                row=i, column=j).value = copied_data[countRow][countCol]
            countCol += 1
        countRow += 1
    to_wb.save(input_filepath)



def df_convert_column_to_type(df: pd.DataFrame, column_name: str, column_type: str) -> pd.DataFrame:
    """Converts the column to the given type.

    Args:
        df (pd.DataFrame): Input dataframe.
        column_name (str): Name of the column.
        column_type (str): Type of the column.
    
    Returns:
        df (pd.DataFrame): Converted dataframe.
    
    Examples:
        >>> df_convert_column_to_type(df, 'column_name', 'str')
        df
    """

    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if str(column_type).lower() in ["string", "str"]:
            df[column_name] = df[column_name].astype('str')
        elif str(column_type).lower() in ["int", "integer"]:
            df[column_name] = df[column_name].astype('int64')
        elif str(column_type).lower() in ["float", "double"]:
            df[column_name] = df[column_name].astype('float64')
        elif str(column_type).lower() in ["date"]:
            df[column_name] = pd.to_datetime(df[column_name], unit='s')
        elif str(column_type).lower() in ["boolean", "bool"]:
            df[column_name] = df[column_name].astype('bool')
        elif str(column_type).lower() in ["complex"]:
            df[column_name] = df[column_name].astype('complex128')
        elif str(column_type).lower() in ["bytes"]:
            df[column_name] = df[column_name].astype('bytes')
        else:
            try:
                df[column_name] = df[column_name].astype(column_type)
            except:
                raise Exception("Please pass correct column type")

        data = df
    return data



def excel_group_by_column_values_n_split(df: pd.DataFrame, column_name: str, output_folder: str, show_output: bool = False) -> None:
    """Groups the dataframe by the given column and splits the dataframe into multiple excel files.

    Args:
        df (pd.DataFrame): Input dataframe.
        column_name (str): Name of the column.
        output_folder (str): Output folder path.
        show_output (bool): If True, shows the output. Default is False.
    
    Examples:
        >>> excel_group_by_column_values_n_split_2(df, 'column_name', 'C:\\Users\\abc\\Desktop')
    """

    # Import Section
    import pandas as pd
    from pathlib import Path
    import os

    # Code Section
    if not column_name:
        raise Exception("Please provide the column name to split.")

    grouped_df = df.groupby(column_name)

    for i in grouped_df:
        file_path = os.path.join(output_folder, str(i[0]) + ".xlsx")
        file_path = Path(file_path)
        grouped_df.get_group(i[0]).to_excel(file_path, index=False)

    if show_output:
        os.startfile(output_folder)


def excel_if_value_exists(df: pd.DataFrame, column:str, value:str) -> bool:
    """Checks if the given value exists in the given column.

    Args:
        df (pd.DataFrame): Input dataframe.
        column (str): Name of the column.
        value (str): Value to be searched.
    
    Returns:
        status (bool): True if value exists, else False.
    
    Examples:
        >>> excel_if_value_exists(df, 'column_name', 'value')
        True
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not value:
        raise Exception("Please provide the value to be searched")

    if not isinstance(cols, list):
        cols = [cols]

    if cols:
        if value in df[cols].values:
            return True
    else:
        if value in df.values:
            return True


def excel_apply_template_format(raw_data_filepath:Union[str, WindowsPath], raw_data_sheetname: str = "Sheet1", template_filepath:Union[str, WindowsPath] = "", template_sheetname: str = "Sheet1", output_folder:Union[str, WindowsPath] = "", output_filename: str = "") -> None: 
    """
    Applies the template format to the excel raw data
    
    Args:
        raw_data_filepath (Union[str, WindowsPath]): The raw data filepath.
        raw_data_sheetname (str, optional): The raw data sheetname. Defaults to "Sheet1".
        template_filepath (Union[str, WindowsPath], optional): The template filepath. Defaults to "".
        template_sheetname (str, optional): The template sheetname. Defaults to "Sheet1".
        output_folder (Union[str, WindowsPath], optional): The output folder. Defaults to "".
        output_filename (str, optional): The output filename. Defaults to "".

    Returns:
        None
    
    Examples:
        >>> excel_apply_template_format_2(raw_data_filepath='my_autopylot\raw_data\raw_data.xlsx', input_sheetname='Sheet1', input_template_filepath='my_autopylot\raw_data\template.xlsx', input_template_sheetname='Sheet1', output_folder='my_autopylot\raw_data', output_filename='raw_data.xlsx')
        None
         
    """

    # Import Section
    import pandas as pd
    from win32com.client import Dispatch
    import os
    
    # Validation section
    if not os.path.isfile(raw_data_filepath):
        raise FileNotFoundError(f'File not found: {raw_data_filepath}')
    
    if not os.path.isfile(template_filepath):
        raise FileNotFoundError(f'File not found: {template_filepath}')

    if output_folder == "" and output_filename == "":
        output_file_path = raw_data_filepath

        #append the output file name with '_formatted'
        output_file_path = os.path.splitext(output_file_path)[0] + '_formatted' + os.path.splitext(output_file_path)[1]

    else:
        output_file_path = os.path.join(output_folder, output_filename)

    if os.path.isfile(output_file_path):
        os.remove(output_file_path)
        

    # Code Section
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.ScreenUpdating = False

    # copy from source
    source = excel.Workbooks.Open(raw_data_filepath)

    #select worksheet
    source_sheet = source.Worksheets(raw_data_sheetname)

    #activate worksheet
    source_sheet.Activate()

    # excel.Range("A2:D4").Select()
    excel.ActiveSheet.UsedRange.Select()
    excel.Selection.Copy()

    # paste (appended) to target
    target = excel.Workbooks.Open(template_filepath)

    #select worksheet

    ws = target.Worksheets(template_sheetname)

    destrange = "A"+str(1) #+":C9"
    
    excel.Range(destrange).Select()
    excel.Selection.PasteSpecial()

    # save and close
    shell = Dispatch("WScript.Shell")
    shell.SendKeys("{ENTER}", 0)

    excel.Selection.Columns.AutoFit()

    excel.ScreenUpdating = True
    
    #save as excel format
    ws.SaveAs(output_file_path, 51) # 51 = excel format
    target.Close()

    source.Close()
    excel.Quit()


def excel_concat_all_sheets_of_given_excel(excel_file_path: Union[str, WindowsPath], sheet_names_as_list=None) -> pd.DataFrame:
    """ Concat all sheets of given excel file

    Args:
        excel_file_path (Union[str, WindowsPath]): The excel file path.
        sheet_names_as_list (list, optional): The sheet names as list. Defaults to None.
    
    Returns:
        data (DataFrame): The data.
    
    Examples:
        >>> excel_concat_all_sheets_of_given_excel(excel_file_path='my_autopylot\raw_data\raw_data.xlsx', sheet_names_as_list=['Sheet1', 'Sheet2'])
        data 
    """

    # Import Section
    import pandas as pd
    
    # Code Section
    if not excel_file_path:
        raise Exception("Please pass excel file path")

    if isinstance(excel_file_path, str):
        if sheet_names_as_list == None:
            data = pd.read_excel(excel_file_path, sheet_name=None)
        elif isinstance(sheet_names_as_list, list):
            data = pd.read_excel(
                excel_file_path, sheet_name=sheet_names_as_list)
        data = pd.concat(data, ignore_index=True)
    else:
        raise Exception("Please pass sheet names as list")

    return data


################################################################################################




